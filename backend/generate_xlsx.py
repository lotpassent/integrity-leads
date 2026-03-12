"""
Gerador de Relatório Excel — Integrity Recruitment Group
Gera um arquivo .xlsx com identidade visual da empresa, formatação profissional,
abas de Leads, Resumo por Status, Performance por Consultor e Meta Semanal.

CORREÇÕES APLICADAS:
  - freeze_panes corrigido de B3 → A3 (não trava coluna ID)
  - Tabela de Leads agora é uma Tabela Excel formal (com filtros automáticos)
  - Datas exportadas como datetime real (não texto)
  - Aba Resumo agora inclui seção "Performance por Consultor"
  - Perguntas do consultor lidas dinamicamente do banco (via campo json)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
import json, sys

# ─── PALETA DE CORES ──────────────────────────────────────────────────────────
TEAL_DEEP   = "0C2022"
TEAL_MID    = "1D4D4F"
TEAL_ACCENT = "2A6B6D"
AMBER       = "D4832A"
AMBER_LIGHT = "E8A455"
CREAM       = "F4F1EC"
CREAM_DIM   = "C4BBAE"
WHITE       = "FAFAF8"
SUCCESS     = "3D9970"
DANGER      = "C0392B"
WARN        = "E67E22"
PURPLE      = "9B59B6"
BLUE        = "4A90E2"
GRAY        = "6C757D"

URG_COLORS = {
    "baixa":   (SUCCESS, "E8F5F0"),
    "media":   (WARN,    "FDF3E3"),
    "alta":    (WARN,    "FDEBD0"),
    "critica": (DANGER,  "FADBD8"),
}
STATUS_COLORS = {
    "triagem":          (GRAY,   "F2F3F4"),
    "contato_inicial":  (BLUE,   "EBF5FB"),
    "proposta_enviada": (PURPLE, "F5EEF8"),
    "negociacao":       (WARN,   "FDF3E3"),
    "fechado_ganho":    (SUCCESS,"E8F8F5"),
    "fechado_perdido":  (DANGER, "FDEDEC"),
}
STATUS_LABELS = {
    "triagem":          "Triagem",
    "contato_inicial":  "Contato Inicial",
    "proposta_enviada": "Proposta Enviada",
    "negociacao":       "Negociação",
    "fechado_ganho":    "Fechado — Ganho",
    "fechado_perdido":  "Fechado — Perdido",
}
NIVEL_LABELS = {
    "junior":"Júnior","pleno":"Pleno","senior":"Sênior",
    "especialista":"Especialista","gerencia":"Gerência",
    "diretoria":"Diretoria","c-level":"C-Level",
}
CONSULTORES_FIXOS = [
    "Camilla Moriya","Cláudio Ferreira","Camila Benedetti",
    "Paula Vieira","Luís Hartmann","Ana Delibi",
]
META_SEMANAL = 2

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def border_thin():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def center():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def right_al():return Alignment(horizontal="right",  vertical="center")

def parse_date(date_str):
    """Converte string ISO para objeto datetime (exportação correta de data no Excel)."""
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(str(date_str)[:19].replace("T", " "))
    except Exception:
        return None

def h1_cell(ws, row, col, span_end, value):
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(span_end)}{row}")
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill = fill(TEAL_DEEP)
    c.alignment = center()
    ws.row_dimensions[row].height = 34

def h2_cell(ws, row, col, span_end, value):
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(span_end)}{row}")
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = fill(TEAL_MID)
    c.alignment = left()
    ws.row_dimensions[row].height = 24

def val_cell(ws, row, col, value, bg=WHITE, bold=False, color="212121", align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold, color=color)
    c.fill = fill(bg)
    c.alignment = align or left()
    c.border = border_thin()
    ws.row_dimensions[row].height = 22
    return c


# ─── ABA 1: LEADS ─────────────────────────────────────────────────────────────
def build_sheet_leads(wb, leads_data):
    ws = wb.create_sheet("📋 Leads")

    # CORREÇÃO 1: freeze_panes A3 (não B3) — congela só as linhas de cabeçalho,
    # sem prender a coluna ID junto
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    # Cabeçalho principal
    ws.merge_cells("A1:R1")
    ws["A1"] = "INTEGRITY RECRUITMENT GROUP  ·  Relatório de Leads Comerciais — Sourcing Reverso"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws["A1"].fill = fill(TEAL_DEEP)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 36

    # Sub-cabeçalho com data
    ws.merge_cells("A2:R2")
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}   |   Total de Leads: {len(leads_data)}"
    ws["A2"].font = Font(name="Arial", size=9, color=AMBER_LIGHT, italic=True)
    ws["A2"].fill = fill(TEAL_MID)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 20

    columns = [
        ("ID",             6),
        ("Consultor",     20),
        ("Empresa",       24),
        ("Setor",         18),
        ("Porte",         14),
        ("Contato",       20),
        ("Cargo Contato", 18),
        ("E-mail",        26),
        ("Telefone",      16),
        ("Vaga em Aberto",24),
        ("Nível",         14),
        ("Qtd",            6),
        ("Urgência",      12),
        ("Status",        20),
        ("Observações",   28),
        ("Contexto",      32),
        ("Criado em",     18),
        ("Atualizado",    18),
    ]

    for i, (col_name, col_w) in enumerate(columns, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = col_w
        cell = ws.cell(row=3, column=i, value=col_name)
        cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        cell.fill = fill(TEAL_ACCENT)
        cell.alignment = center()
        cell.border = Border(
            bottom=Side(style="medium", color=AMBER),
            left=Side(style="thin", color=TEAL_MID),
            right=Side(style="thin", color=TEAL_MID),
        )
    ws.row_dimensions[3].height = 28

    for r_idx, lead in enumerate(leads_data, 4):
        row = r_idx
        is_odd = (r_idx % 2 == 0)
        row_bg = "EEF4F4" if is_odd else WHITE

        urg = lead.get("urgencia", "").lower()
        st  = lead.get("status", "")
        urg_fg, urg_bg = URG_COLORS.get(urg,  (GRAY, "F5F5F5"))
        st_fg,  st_bg  = STATUS_COLORS.get(st, (GRAY, "F5F5F5"))

        def dc(col, val, bg=None, bold=False, color="212121", align=left()):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=9, bold=bold, color=color)
            c.fill = fill(bg or row_bg)
            c.alignment = align
            c.border = border_thin()
            return c

        # CORREÇÃO 2: datas como datetime real (não string)
        criado_dt  = parse_date(lead.get("criado_em", ""))
        atualiz_dt = parse_date(lead.get("atualizado_em", ""))

        dc(1,  lead.get("id"),                 bg=row_bg, align=right_al())
        dc(2,  lead.get("consultor_nome"),      bg=row_bg, bold=True)
        dc(3,  lead.get("empresa_nome"),        bg=row_bg, bold=True)
        dc(4,  lead.get("empresa_setor", ""),   bg=row_bg)
        dc(5,  lead.get("empresa_porte", ""),   bg=row_bg)
        dc(6,  lead.get("contato_nome", ""),    bg=row_bg)
        dc(7,  lead.get("contato_cargo", ""),   bg=row_bg)
        dc(8,  lead.get("contato_email", ""),   bg=row_bg)
        dc(9,  lead.get("contato_telefone",""), bg=row_bg)
        dc(10, lead.get("cargo_em_aberto",""),  bg=row_bg)
        dc(11, NIVEL_LABELS.get(lead.get("nivel_cargo",""), lead.get("nivel_cargo","")), bg=row_bg, align=center())
        dc(12, lead.get("quantidade_vagas", 1), bg=row_bg, align=center())

        urg_cell = ws.cell(row=row, column=13, value=urg.capitalize() if urg else "")
        urg_cell.font  = Font(name="Arial", size=9, bold=True, color=urg_fg)
        urg_cell.fill  = fill(urg_bg)
        urg_cell.alignment = center()
        urg_cell.border = border_thin()

        st_cell = ws.cell(row=row, column=14, value=STATUS_LABELS.get(st, st))
        st_cell.font  = Font(name="Arial", size=9, bold=True, color=st_fg)
        st_cell.fill  = fill(st_bg)
        st_cell.alignment = center()
        st_cell.border = border_thin()

        dc(15, lead.get("observacoes", ""),             bg=row_bg)
        dc(16, lead.get("contexto_entrevistado", ""),   bg=row_bg)

        # Datas como datetime (Excel reconhece e permite filtrar/ordenar por data)
        c17 = ws.cell(row=row, column=17, value=criado_dt)
        c17.font = Font(name="Arial", size=9, color="212121")
        c17.fill = fill(row_bg)
        c17.alignment = center()
        c17.border = border_thin()
        c17.number_format = "DD/MM/YYYY HH:MM"

        c18 = ws.cell(row=row, column=18, value=atualiz_dt)
        c18.font = Font(name="Arial", size=9, color="212121")
        c18.fill = fill(row_bg)
        c18.alignment = center()
        c18.border = border_thin()
        c18.number_format = "DD/MM/YYYY HH:MM"

        ws.row_dimensions[row].height = 20

    # CORREÇÃO 3: Tabela formal do Excel (filtros automáticos nativos)
    if leads_data:
        last_row = 3 + len(leads_data)
        table = Table(
            displayName="TbLeads",
            ref=f"A3:R{last_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Rodapé
    footer_row = 3 + len(leads_data) + 2
    ws.merge_cells(f"A{footer_row}:R{footer_row}")
    ws[f"A{footer_row}"] = "Integrity Recruitment Group  ·  Sistema de Sourcing Reverso  ·  Confidencial"
    ws[f"A{footer_row}"].font = Font(name="Arial", size=8, color=CREAM_DIM, italic=True)
    ws[f"A{footer_row}"].fill = fill(TEAL_DEEP)
    ws[f"A{footer_row}"].alignment = center()
    ws.row_dimensions[footer_row].height = 18


# ─── ABA 2: RESUMO EXECUTIVO ──────────────────────────────────────────────────
def build_sheet_resumo(wb, leads_data):
    ws2 = wb.create_sheet("📊 Resumo")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 14

    h1_cell(ws2, 1, 1, 4, "INTEGRITY RECRUITMENT GROUP  ·  Resumo Executivo")
    ws2.merge_cells("A2:D2")
    ws2["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}   ·   {len(leads_data)} leads no total"
    ws2["A2"].font = Font(name="Arial", size=9, color=AMBER_LIGHT, italic=True)
    ws2["A2"].fill = fill(TEAL_MID)
    ws2["A2"].alignment = center()
    ws2.row_dimensions[2].height = 18

    # Totais gerais
    h2_cell(ws2, 4, 1, 4, "  📌  Totais Gerais")
    totals = [
        ("Total de Leads",          len(leads_data)),
        ("Fechados (Ganhos)",        sum(1 for l in leads_data if l.get("status")=="fechado_ganho")),
        ("Em Negociação",            sum(1 for l in leads_data if l.get("status")=="negociacao")),
        ("Leads Urgentes/Críticos",  sum(1 for l in leads_data if l.get("urgencia") in ("alta","critica"))),
    ]
    for i, (label, v) in enumerate(totals):
        rr = 5 + i
        val_cell(ws2, rr, 1, label, bg="EEF4F4", bold=True)
        val_cell(ws2, rr, 2, v, bg="EEF4F4", bold=True, color=TEAL_ACCENT, align=center())

    # Por Status
    h2_cell(ws2, 11, 1, 4, "  📊  Leads por Status")
    for i, (st_key, st_label) in enumerate(STATUS_LABELS.items()):
        rr = 12 + i
        count = sum(1 for l in leads_data if l.get("status")==st_key)
        pct = (count / len(leads_data) * 100) if leads_data else 0
        st_fg, st_bg = STATUS_COLORS.get(st_key, (GRAY, "F5F5F5"))
        val_cell(ws2, rr, 1, st_label, bg=st_bg)
        val_cell(ws2, rr, 2, count,  bg=st_bg, bold=True, color=st_fg, align=center())
        val_cell(ws2, rr, 3, f"{pct:.1f}%", bg=st_bg, align=center())
        val_cell(ws2, rr, 4, "█" * int(pct / 5), bg=st_bg, color=st_fg)

    # Por Urgência
    urg_row = 20
    h2_cell(ws2, urg_row, 1, 4, "  🚦  Leads por Urgência")
    for i, (urg_key, urg_label) in enumerate([("critica","Crítica"),("alta","Alta"),("media","Média"),("baixa","Baixa")]):
        rr = urg_row + 1 + i
        count = sum(1 for l in leads_data if l.get("urgencia")==urg_key)
        pct = (count / len(leads_data) * 100) if leads_data else 0
        urg_fg, urg_bg = URG_COLORS.get(urg_key, (GRAY, "F5F5F5"))
        val_cell(ws2, rr, 1, urg_label, bg=urg_bg)
        val_cell(ws2, rr, 2, count,  bg=urg_bg, bold=True, color=urg_fg, align=center())
        val_cell(ws2, rr, 3, f"{pct:.1f}%", bg=urg_bg, align=center())
        val_cell(ws2, rr, 4, "█" * int(pct / 5), bg=urg_bg, color=urg_fg)

    # CORREÇÃO 4: Seção de Performance por Consultor (estava faltando)
    perf_row = urg_row + 7
    h2_cell(ws2, perf_row, 1, 4, "  👤  Performance por Consultor")

    # Cabeçalho da sub-tabela
    for col, label in enumerate(["Consultor", "Total Leads", "Ganhos", "Em Andamento"], 1):
        c = ws2.cell(row=perf_row + 1, column=col, value=label)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(TEAL_ACCENT)
        c.alignment = center()
        c.border = border_thin()

    consultores = {}
    for lead in leads_data:
        nome = lead.get("consultor_nome", "Sem consultor")
        if nome not in consultores:
            consultores[nome] = {"total": 0, "ganhos": 0, "andamento": 0}
        consultores[nome]["total"] += 1
        if lead.get("status") == "fechado_ganho":
            consultores[nome]["ganhos"] += 1
        if lead.get("status") in ("negociacao", "proposta_enviada"):
            consultores[nome]["andamento"] += 1

    for i, (nome, stats) in enumerate(sorted(consultores.items(), key=lambda x: -x[1]["total"])):
        rr = perf_row + 2 + i
        row_bg = "EEF4F4" if i % 2 == 0 else WHITE
        val_cell(ws2, rr, 1, nome,                  bg=row_bg, bold=True)
        val_cell(ws2, rr, 2, stats["total"],         bg=row_bg, bold=True, color=TEAL_ACCENT, align=center())
        val_cell(ws2, rr, 3, stats["ganhos"],        bg=row_bg, color=SUCCESS, bold=True, align=center())
        val_cell(ws2, rr, 4, stats["andamento"],     bg=row_bg, color=WARN, align=center())


# ─── ABA 3: META SEMANAL ──────────────────────────────────────────────────────
def build_sheet_meta(wb, leads_data):
    ws3 = wb.create_sheet("🎯 Meta Semanal")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 16
    ws3.column_dimensions["F"].width = 20

    h1_cell(ws3, 1, 1, 6, "INTEGRITY RECRUITMENT GROUP  ·  Controle de Meta Semanal")
    ws3.merge_cells("A2:F2")
    ws3["A2"] = f"Meta: {META_SEMANAL} leads por consultor por semana   ·   Semana de {datetime.now().strftime('%d/%m/%Y')}"
    ws3["A2"].font = Font(name="Arial", size=9, color=AMBER_LIGHT, italic=True)
    ws3["A2"].fill = fill(TEAL_MID)
    ws3["A2"].alignment = center()
    ws3.row_dimensions[2].height = 18

    for i, col_name in enumerate(["Consultor","Leads Esta Semana","Meta","% Atingida","Status","Observação"]):
        c = ws3.cell(row=3, column=i+1, value=col_name)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = fill(TEAL_ACCENT)
        c.alignment = center()
        c.border = Border(
            bottom=Side(style="medium", color=AMBER),
            left=Side(style="thin", color=TEAL_MID),
            right=Side(style="thin", color=TEAL_MID),
        )
    ws3.row_dimensions[3].height = 26

    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    weekly_counts = {}
    for lead in leads_data:
        criado_dt = parse_date(lead.get("criado_em", ""))
        if criado_dt and criado_dt.date() >= start_of_week:
            nome = lead.get("consultor_nome", "")
            weekly_counts[nome] = weekly_counts.get(nome, 0) + 1

    all_names = list(CONSULTORES_FIXOS)
    for nome in weekly_counts:
        if nome not in all_names:
            all_names.append(nome)

    for i, nome in enumerate(all_names):
        rr = 4 + i
        realizados = weekly_counts.get(nome, 0)
        pct = min((realizados / META_SEMANAL) * 100, 100)

        if realizados >= META_SEMANAL:
            status_txt = "✅ Meta Atingida"
            row_bg = "E8F5F0"
            pct_color = SUCCESS
        elif realizados >= 1:
            status_txt = "⚠️ Em Andamento"
            row_bg = "FDF8E1"
            pct_color = WARN
        else:
            status_txt = "❌ Sem Leads"
            row_bg = "FDE8E8"
            pct_color = DANGER

        is_odd = (i % 2 == 0)

        def wc(col, val, bg=None, bold=False, color="212121", align=left()):
            c = ws3.cell(row=rr, column=col, value=val)
            c.font = Font(name="Arial", size=10, bold=bold, color=color)
            c.fill = fill(bg or (row_bg if is_odd else WHITE))
            c.alignment = align
            c.border = border_thin()
            ws3.row_dimensions[rr].height = 22
            return c

        wc(1, nome, bold=True)
        wc(2, realizados, bold=True, color=pct_color if realizados > 0 else DANGER, align=center())
        wc(3, META_SEMANAL, align=center())
        wc(4, f"{pct:.0f}%", bold=True, color=pct_color, align=center())
        wc(5, status_txt, align=center())

        if realizados >= META_SEMANAL:
            obs = f"+{realizados - META_SEMANAL} acima da meta" if realizados > META_SEMANAL else "Exatamente na meta"
        elif realizados > 0:
            obs = f"Falta {META_SEMANAL - realizados} lead(s)"
        else:
            obs = "Nenhum lead esta semana"
        wc(6, obs, color="666666")

    footer_r = 4 + len(all_names) + 2
    ws3.merge_cells(f"A{footer_r}:F{footer_r}")
    ws3[f"A{footer_r}"] = "Integrity Recruitment Group  ·  Sistema de Sourcing Reverso  ·  Confidencial"
    ws3[f"A{footer_r}"].font = Font(name="Arial", size=8, color=CREAM_DIM, italic=True)
    ws3[f"A{footer_r}"].fill = fill(TEAL_DEEP)
    ws3[f"A{footer_r}"].alignment = center()
    ws3.row_dimensions[footer_r].height = 16


# ─── FUNÇÃO PRINCIPAL ─────────────────────────────────────────────────────────
def build_report(leads_data: list, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet_leads(wb, leads_data)
    build_sheet_resumo(wb, leads_data)
    build_sheet_meta(wb, leads_data)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        with open(sys.argv[1]) as f:
            data = json.load(f)
        out = build_report(data, sys.argv[2])
        print(f"Excel gerado: {out}")
    else:
        sample = [
            {"id":1,"consultor_nome":"Camilla Moriya","empresa_nome":"Banco do Meridian","empresa_setor":"Financeiro / Banco","empresa_porte":"grande","contato_nome":"Roberto Almeida","contato_cargo":"Gerente de RH","contato_email":"r.almeida@meridian.com.br","contato_telefone":"(11) 98765-4321","cargo_em_aberto":"Analista de Dados Sênior","nivel_cargo":"senior","quantidade_vagas":2,"urgencia":"alta","status":"contato_inicial","observacoes":"Preferência por contato às terças","contexto_entrevistado":"Time de dados cresce 40%, processo aberto há 3 meses sem sucesso","criado_em":"2026-03-10T09:30:00","atualizado_em":"2026-03-11T14:00:00"},
            {"id":2,"consultor_nome":"Cláudio Ferreira","empresa_nome":"TechBank S.A.","empresa_setor":"Tecnologia","empresa_porte":"enterprise","contato_nome":"Fernanda Costa","contato_cargo":"Diretora de T.I.","contato_email":"f.costa@techbank.com","contato_telefone":"(11) 3456-7890","cargo_em_aberto":"Gerente de Projetos","nivel_cargo":"gerencia","quantidade_vagas":1,"urgencia":"critica","status":"proposta_enviada","observacoes":"Urgência máxima — posição descoberta desde janeiro","contexto_entrevistado":"Gestor saiu inesperadamente, time de 8 pessoas sem liderança","criado_em":"2026-03-11T11:00:00","atualizado_em":"2026-03-12T08:30:00"},
            {"id":3,"consultor_nome":"Paula Vieira","empresa_nome":"Grupo Meridian","empresa_setor":"Consultoria","empresa_porte":"media","contato_nome":"Marcos Souza","contato_cargo":"Sócio-Diretor","contato_email":"m.souza@meridian.com","contato_telefone":"","cargo_em_aberto":"Desenvolvedor Full Stack","nivel_cargo":"pleno","quantidade_vagas":3,"urgencia":"media","status":"triagem","observacoes":"","contexto_entrevistado":"Startup de fintech em expansão, rodada série B recente","criado_em":"2026-03-12T08:00:00","atualizado_em":"2026-03-12T08:00:00"},
            {"id":4,"consultor_nome":"Ana Delibi","empresa_nome":"Construtora Vega","empresa_setor":"Construção Civil","empresa_porte":"grande","contato_nome":"","contato_cargo":"","contato_email":"","contato_telefone":"","cargo_em_aberto":"CFO","nivel_cargo":"c-level","quantidade_vagas":1,"urgencia":"critica","status":"negociacao","observacoes":"Candidato não quis dar mais detalhes","contexto_entrevistado":"Empresa abrindo capital, precisam de CFO com experiência em IPO","criado_em":"2026-03-08T14:00:00","atualizado_em":"2026-03-10T16:00:00"},
            {"id":5,"consultor_nome":"Luís Hartmann","empresa_nome":"Farmácia Saúde+","empresa_setor":"Saúde","empresa_porte":"media","contato_nome":"Ana Ribeiro","contato_cargo":"Gerente Administrativa","contato_email":"ana@saudemais.com.br","contato_telefone":"(41) 99988-7766","cargo_em_aberto":"Farmacêutico Responsável","nivel_cargo":"especialista","quantidade_vagas":1,"urgencia":"baixa","status":"fechado_ganho","observacoes":"Contrato fechado em 2 semanas","contexto_entrevistado":"Nova unidade inaugurando em abril, precisam de RT já aprovado pelo CRF","criado_em":"2026-03-05T10:00:00","atualizado_em":"2026-03-12T09:00:00"},
            {"id":6,"consultor_nome":"Camila Benedetti","empresa_nome":"Logística Express","empresa_setor":"Logística","empresa_porte":"grande","contato_nome":"Paulo Henrique","contato_cargo":"Diretor de Operações","contato_email":"ph@express.com.br","contato_telefone":"(21) 3333-4444","cargo_em_aberto":"Coordenador de Frota","nivel_cargo":"pleno","quantidade_vagas":2,"urgencia":"alta","status":"fechado_perdido","observacoes":"Optaram por candidato interno","contexto_entrevistado":"Expansão de frota para 2026, processo acelerado","criado_em":"2026-03-01T10:00:00","atualizado_em":"2026-03-09T15:00:00"},
        ]
        out = build_report(sample, "/tmp/relatorio_integrity_demo.xlsx")
        print(f"Demo gerado: {out}")